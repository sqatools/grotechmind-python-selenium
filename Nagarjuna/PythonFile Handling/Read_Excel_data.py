import openpyxl
"""
pip install openpyxl
"""

def read_excel_file(filepath, row, col):
    wb_object = openpyxl.load_workbook(filepath)
    sheet_obj = wb_object.active
    cell_object = sheet_obj.cell(row=row, column=col)
    return cell_object.value

file_path = r"D:\Python\GitRepo\grotechmind-python-selenium\Nagarjuna\PythonFile Handling\Test_Data.xlsx"
val = read_excel_file(file_path, 1, 1)
print(val)
"""
print("@"*20)
for i in range(1, 4):
    val = read_excel_file(file_path, i, 1)
    print(val)

print("%"*30)
"""
def get_max_row_col(filepath):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active
    max_row = sheet.max_row
    max_column = sheet.max_column
    return max_row, max_column, sheet

output = get_max_row_col(file_path)
print(output)

sheet = output[2]
max_row = output[0]
max_col = output[1]

for i in range(1, max_row+1):
    for j in range(1, max_col+1):
        print(sheet.cell(row=i, column=j).value, end=" ")
    print()

