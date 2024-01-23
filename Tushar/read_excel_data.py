import openpyxl
"""
def read_excel_data(file_name,row,col):
    wb_obj=openpyxl.load_workbook(file_path)
    sheet_obj=wb_obj.active
    cell_obj=sheet_obj.cell(row=row,column=col)
    return cell_obj.value

file_path=r"D:\Python_Automation\Test_data.xlsx"
val=read_excel_data(file_path,1,1)
print(val)


for i in range(1, 6):
    val = read_excel_data(file_path, i, 1)
    print(val)
"""
"""
def read_excel_data(file_path,row,col):
    wb=openpyxl.load_workbook(file_path)
    sheet=wb.active
    cell=sheet.cell(row=row,column=col)
    return cell.value
file_path=r"D:\Python_Automation\Test_data.xlsx"
result=read_excel_data(file_path,1,1)
print(result)

for i in range(1,6):
    val=read_excel_data(file_path,i,1)
    print(val)
"""

# def read_excel_data(file_path,row,col):
#     wb=openpyxl.load_workbook(file_path)
#     sheet=wb.active
#     cell=sheet.cell(row=row,column=col)
#     return cell.value
# file_path=r"C:\Users\Lenovo\Desktop\EMP.xlsx"
# result=read_excel_data(file_path,1,1)
# print(result)
#
# for i in range (1,12):
#     result=read_excel_data(file_path,i,1)
#     print(result)

# def read_excel_data(file_path,row,col):
#     wb=openpyxl.load_workbook(file_path)
#     sheet=wb.active
#     cell=sheet.cell(row=row,column=col)
#     return cell.value
# file_path=r"C:\Users\Lenovo\Desktop\EMP.xlsx"
# result=read_excel_data(file_path,1,1)
# print(result)
#
# for i in range(1,6):
#     result=read_excel_data(file_path,i,1)
#     print(result)



def get_max_row_col(filepath):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active
    max_row = sheet.max_row
    max_column = sheet.max_column
    return max_row, max_column, sheet
file_path=r"C:\Users\Lenovo\Desktop\EMP.xlsx"
output = get_max_row_col(file_path)
print(output)

sheet = output[2]
max_row = output[0]
max_col = output[1]

for i in range(2, max_row+1):
    for j in range(1, max_col+1):
        print(sheet.cell(row=i, column=j).value, end=" | ")
    print()


























































