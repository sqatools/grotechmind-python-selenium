import openpyxl

def read_excel_file(filepath, row, col):
    wb_object = openpyxl.load_workbook(filepath)
    sheet_obj = wb_object.active
    cell_obj = sheet_obj.cell(row=row,column=col)
    return cell_obj.value
filepath = r"E:\Training\grotechmind-python-selenium\Banupriya\Python FileHandling\FileTestData1.xlsx"
#val = read_excel_file(filepath,2,2 )
#print(val)

for i in range(1,4):
    val = read_excel_file(filepath, i,4)
    print(val)

""" 
o/p
A
B
C
"""

def get_max_row_col(filepath):
    wb = openpyxl.load_workbook(filepath)
    sheet= wb.active
    max_row = sheet.max_row
    max_col = sheet.max_column
    return max_row, max_col, sheet

output = get_max_row_col(filepath)
print(output)

sheet = output[2]
max_row = output[0]
max_col = output[1]

for i in range(1,max_row+1):
    for j in range(1,max_col+1):
        print(sheet.cell(row=i,column=j).value,end ='|')
    print()

